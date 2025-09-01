"""
Excel report generation module for GCal Extractor.
Creates Excel files with 'totales' and 'detalle' sheets.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List, Dict
import os
import streamlit as st


class ExcelGenerator:
    """Handles Excel report generation with dual sheets."""
    
    def __init__(self):
        self.reports_dir = 'reports'
        # Ensure reports directory exists
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_timestamped_filename(self, year: int, month: int) -> str:
        """Generate timestamped filename for Excel report."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"report_{year:04d}_{month:02d}_{timestamp}.xlsx"
    
    def create_totales_sheet(self, workbook: Workbook, totales_data: List[Dict]):
        """Create 'totales' sheet with calendar, patient name, and total columns."""
        ws = workbook.active
        ws.title = "totales"
        
        # Create DataFrame
        df = pd.DataFrame(totales_data)
        
        # Add headers
        headers = ['calendario', 'nombre', 'total']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_num, row_data in enumerate(totales_data, 2):
            ws.cell(row=row_num, column=1, value=row_data['calendario'])
            ws.cell(row=row_num, column=2, value=row_data['nombre'])
            ws.cell(row=row_num, column=3, value=row_data['total'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_detalle_sheet(self, workbook: Workbook, detalle_data: Dict[str, List[Dict]]):
        """Create 'detalle' sheet with patient columns and consultation dates."""
        ws = workbook.create_sheet(title="detalle")
        
        current_col = 1
        
        for calendar_name, patients in detalle_data.items():
            if not patients:
                continue
            
            # Add calendar name header (spans across patient columns)
            start_col = current_col
            
            # Add patient headers
            patient_cols = []
            for patient in patients:
                patient_name = patient['patient_name']
                
                # Write patient name header
                cell = ws.cell(row=1, column=current_col, value=patient_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                
                patient_cols.append(current_col)
                current_col += 1
            
            # Add calendar name header spanning patient columns
            if len(patient_cols) > 1:
                ws.merge_cells(start_row=2, start_column=start_col, 
                              end_row=2, end_column=current_col-1)
            
            calendar_cell = ws.cell(row=2, column=start_col, value=calendar_name)
            calendar_cell.font = Font(bold=True)
            calendar_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            calendar_cell.alignment = Alignment(horizontal="center")
            
            # Add patient dates
            for i, patient in enumerate(patients):
                col = patient_cols[i]
                dates = patient['dates']
                
                for row_num, date in enumerate(dates, 3):
                    ws.cell(row=row_num, column=col, value=date)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_excel_report(self, totales_data: List[Dict], detalle_data: Dict[str, List[Dict]], 
                            year: int, month: int) -> str:
        """
        Generate complete Excel report with both sheets.
        
        Returns the filepath of the generated Excel file.
        """
        try:
            # Create workbook
            workbook = Workbook()
            
            # Create totales sheet
            self.create_totales_sheet(workbook, totales_data)
            
            # Create detalle sheet
            self.create_detalle_sheet(workbook, detalle_data)
            
            # Generate filename and save
            filename = self.generate_timestamped_filename(year, month)
            filepath = os.path.join(self.reports_dir, filename)
            
            workbook.save(filepath)
            
            return filepath
            
        except Exception as e:
            st.error(f"Error generating Excel report: {e}")
            return None
    
    def get_report_summary(self, totales_data: List[Dict], detalle_data: Dict[str, List[Dict]]) -> Dict:
        """Generate summary statistics for the report."""
        total_patients = len(totales_data)
        total_sessions = sum(row['total'] for row in totales_data)
        calendars_count = len(detalle_data)
        
        # Calculate sessions per calendar
        calendar_stats = {}
        for calendar_name, patients in detalle_data.items():
            calendar_sessions = sum(len(patient['dates']) for patient in patients)
            calendar_stats[calendar_name] = {
                'patients': len(patients),
                'sessions': calendar_sessions
            }
        
        return {
            'total_patients': total_patients,
            'total_sessions': total_sessions,
            'calendars_count': calendars_count,
            'calendar_stats': calendar_stats
        }
